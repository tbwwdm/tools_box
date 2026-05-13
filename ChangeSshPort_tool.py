import sys
import os
import logging
import socket
from datetime import datetime

from PySide6.QtWidgets import (
    QApplication, QWidget, QPushButton, QLabel, QLineEdit,
    QTextEdit, QFileDialog, QVBoxLayout, QHBoxLayout,
    QMessageBox, QCheckBox, QComboBox, QGroupBox
)
from PySide6.QtCore import QObject, Signal, QThread
from PySide6.QtGui import QFont

import paramiko
from openpyxl import load_workbook


logger = logging.getLogger(__name__)


# ================= SSHD 配置应用 =================
def apply_sshd_config(ssh, options: dict):
    commands = [
        "cp /etc/ssh/sshd_config /etc/ssh/sshd_config.bak"
    ]

    for key, value in options.items():
        commands.append(f"sed -i '/^#\\?{key} /d' /etc/ssh/sshd_config")
        commands.append(f"echo '{key} {value}' >> /etc/ssh/sshd_config")

    commands.append("systemctl restart sshd")

    for cmd in commands:
        stdin, stdout, stderr = ssh.exec_command(cmd)
        err = stderr.read().decode().strip()
        if err:
            raise Exception(err)


def modify_sshd(ip, username, password, login_port, sshd_options, log_cb):
    try:
        #log_cb(f"[{ip}] 连接中...")

        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(
            hostname=ip,
            port=int(login_port),
            username=username,
            password=password,
            timeout=10
        )

        apply_sshd_config(ssh, sshd_options)
        ssh.close()

        # 构造更详细日志
        changes = []
        for k, v in sshd_options.items():
            if k == "Port":
                changes.append(f"端口修改为 {v}")
            elif k == "PermitRootLogin":
                if v.lower() == "no":
                    changes.append("禁止 root 用户远程登录")
                else:
                    changes.append("允许 root 用户远程登录")
            else:
                changes.append(f"{k} 设置为 {v}")

        log_cb(f"[{ip}] ✅ sshd 配置修改成功，" + "；".join(changes))

    except (paramiko.SSHException, socket.error, Exception) as e:
        log_cb(f"[{ip}] ❌ 失败: {e}")
        logger.exception(f"{ip} 执行失败")



# ================= Worker =================
class Worker(QObject):
    log_signal = Signal(str)
    finished_signal = Signal()

    def __init__(self, tasks, sshd_options):
        super().__init__()
        self.tasks = tasks
        self.sshd_options = sshd_options

    def run(self):
        for row in self.tasks:
            host = str(row[0]) if len(row) > 0 else ""
            port = int(row[1]) if len(row) > 1 and row[1] is not None else 22
            user = str(row[2]) if len(row) > 2 else ""
            pwd = str(row[3]) if len(row) > 3 else ""

            if not all([host, user, pwd]):
                self.log_signal.emit(f"[{host}] ❌ 数据不完整，跳过")
                continue

            modify_sshd(
                host, user, pwd, port,
                self.sshd_options,
                self.log_signal.emit
            )

        self.finished_signal.emit()


# ================= GUI =================
class SSHPortGUI(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("sshd 批量配置工具")
        self.resize(900, 600)

        self.excel_path = ""
        self.thread = None
        self.worker = None

        self._init_logging()
        self.init_ui()

    def _init_logging(self):
        log_dir = os.path.join(os.path.dirname(__file__), "logs")
        os.makedirs(log_dir, exist_ok=True)
        fh = logging.FileHandler(
            os.path.join(log_dir, f"SSH端口配置_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"),
            encoding="utf-8")
        fh.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        logger.addHandler(fh)
        logger.setLevel(logging.INFO)

    def init_ui(self):
        self.setStyleSheet("""
            SSHPortGUI { background:#f5f6fa; }
            QLineEdit { border:none; border-bottom:1px solid #dfe6e9; padding:6px 4px; background:transparent; font-size:13px; }
            QLineEdit:focus { border-bottom:2px solid #0984e3; }
            QCheckBox { spacing:6px; font-size:13px; }
            QComboBox { border:none; border-bottom:1px solid #dfe6e9; padding:6px 4px; background:transparent; font-size:13px; }
            QComboBox:focus { border-bottom:2px solid #0984e3; }
            QTextEdit { border:1px solid #dfe6e9; border-radius:4px; background:white; font-size:13px; }
        """)
        layout = QVBoxLayout(self)
        layout.setSpacing(8)
        layout.setContentsMargins(28, 24, 28, 24)

        # SSH 端口
        l1 = QHBoxLayout()
        self.cb_port = QCheckBox("修改 SSH 端口")
        l1.addWidget(self.cb_port)
        self.port_input = QLineEdit("22")
        self.port_input.setFixedWidth(70)
        l1.addWidget(self.port_input)
        l1.addStretch()
        layout.addLayout(l1)

        layout.addSpacing(4)

        # Root 登录
        l2 = QHBoxLayout()
        self.cb_root = QCheckBox("禁止 root 远程登录")
        l2.addWidget(self.cb_root)
        l2.addWidget(QLabel("设置"))
        self.root_combo = QComboBox()
        self.root_combo.addItems(["否（允许）", "是（禁止）"])
        l2.addWidget(self.root_combo)
        l2.addStretch()
        layout.addLayout(l2)

        layout.addSpacing(8)

        # 文件
        l3 = QHBoxLayout()
        self.excel_label = QLabel("未选择 Excel 文件")
        l3.addWidget(self.excel_label)
        l3.addStretch()
        excel_btn = QPushButton("选择文件")
        excel_btn.setStyleSheet("QPushButton{background:#0984e3;color:white;padding:7px 20px;border:none;border-radius:4px;font-size:13px;}QPushButton:hover{background:#0873c4;}")
        excel_btn.clicked.connect(self.choose_excel)
        l3.addWidget(excel_btn)
        layout.addLayout(l3)

        layout.addSpacing(12)

        run_btn = QPushButton("开始执行")
        run_btn.setStyleSheet("QPushButton{background:#27ae60;color:white;font-weight:bold;padding:10px 36px;border:none;border-radius:4px;font-size:14px;}QPushButton:hover{background:#219a52;}QPushButton:disabled{background:#b2bec3;}")
        run_btn.clicked.connect(self.start_task)
        layout.addWidget(run_btn)

        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setFont(QFont("Consolas", 10))
        layout.addWidget(self.log_text)

    def log(self, msg):
        self.log_text.append(msg)
        self.log_text.ensureCursorVisible()
        logger.info(msg)

    def choose_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择 Excel 文件", "", "Excel (*.xlsx)"
        )
        if path:
            self.excel_path = path
            self.excel_label.setText(os.path.basename(path))

    def collect_sshd_options(self):
        options = {}

        if self.cb_port.isChecked():
            try:
                port = int(self.port_input.text())
                if not (1 <= port <= 65535):
                    raise ValueError
                options["Port"] = port
            except ValueError:
                raise ValueError("SSH 端口不合法")

        if self.cb_root.isChecked():
            value = self.root_combo.currentText()
            options["PermitRootLogin"] = "no" if value == "是" else "yes"

        if not options:
            raise ValueError("请至少选择一个要修改的配置项")

        return options

    def start_task(self):
        if not self.excel_path:
            QMessageBox.critical(self, "错误", "请先选择 Excel 文件")
            return

        try:
            sshd_options = self.collect_sshd_options()
        except ValueError as e:
            QMessageBox.critical(self, "错误", str(e))
            return

        try:
            wb = load_workbook(self.excel_path)
            sheet = wb.active
            tasks = list(sheet.iter_rows(min_row=2, values_only=True))
        except Exception as e:
            QMessageBox.critical(self, "错误", f"读取 Excel 失败: {e}")
            return

        self.worker = Worker(tasks, sshd_options)
        self.thread = QThread()
        self.worker.moveToThread(self.thread)

        self.thread.started.connect(self.worker.run)
        self.worker.log_signal.connect(self.log)
        self.worker.finished_signal.connect(self.on_finished)
        self.worker.finished_signal.connect(self.thread.quit)

        self.thread.start()
        self.log("=== 任务开始执行 ===")

    def on_finished(self):
        self.log("=== 全部任务执行完成 ===")
        QMessageBox.information(self, "完成", "所有服务器已处理完毕")


# ================= 主入口 =================
if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler("ssh_port_modify.log", encoding="utf-8"),
            logging.StreamHandler(sys.stdout)
        ]
    )
    app = QApplication(sys.argv)
    win = SSHPortGUI()
    win.show()
    sys.exit(app.exec())

