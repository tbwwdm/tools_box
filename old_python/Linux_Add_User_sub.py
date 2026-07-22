import sys
import os
import logging
from datetime import datetime

from PySide6.QtWidgets import (
    QApplication, QWidget, QPushButton, QLabel, QLineEdit,
    QTextEdit, QFileDialog, QVBoxLayout, QHBoxLayout,
    QMessageBox, QCheckBox
)
from PySide6.QtCore import QObject, Signal, QThread
from PySide6.QtGui import QFont

import paramiko
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# ================= SSH 操作 =================
def ssh_create_user(ip, login_user, login_pass, port,
                    new_user, new_pass, use_sudo, log_cb):
    try:
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(ip, int(port), login_user, login_pass, timeout=10)

        prefix = "sudo " if use_sudo else ""

        stdin, stdout, _ = ssh.exec_command(f"id {new_user}")
        if stdout.channel.recv_exit_status() == 0:
            log_cb(f"[{ip}] 用户 {new_user} 已存在，跳过")
            ssh.close()
            return

        ssh.exec_command(f"{prefix}useradd {new_user}")
        ssh.exec_command(f'echo "{new_user}:{new_pass}" | {prefix}chpasswd')

        log_cb(f"[{ip}] ✅ 用户 {new_user} 创建成功")
        ssh.close()

    except Exception as e:
        log_cb(f"[{ip}] ❌ 创建用户失败: {e}")
        logger.exception(ip)


def ssh_delete_user(ip, login_user, login_pass, port,
                    del_user, use_sudo, log_cb):
    try:
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(ip, int(port), login_user, login_pass, timeout=10)

        prefix = "sudo " if use_sudo else ""

        stdin, stdout, _ = ssh.exec_command(f"id {del_user}")
        if stdout.channel.recv_exit_status() != 0:
            log_cb(f"[{ip}] 用户 {del_user} 不存在，跳过")
            ssh.close()
            return

        ssh.exec_command(f"{prefix}userdel -r {del_user}")
        log_cb(f"[{ip}] ✅ 用户 {del_user} 已删除")
        ssh.close()

    except Exception as e:
        log_cb(f"[{ip}] ❌ 删除用户失败: {e}")
        logger.exception(ip)


# ================= Worker =================
class Worker(QObject):
    log_signal = Signal(str)
    finished_signal = Signal()

    def __init__(self, tasks, add_cfg, del_cfg):
        super().__init__()
        self.tasks = tasks
        self.add_cfg = add_cfg
        self.del_cfg = del_cfg

    def run(self):
        for row in self.tasks:
            host = str(row[0]) if len(row) > 0 else ""
            port = int(row[1]) if len(row) > 1 and row[1] is not None else 22
            user = str(row[2]) if len(row) > 2 else ""
            pwd = str(row[3]) if len(row) > 3 else ""

            if not all([host, user, pwd]):
                self.log_signal.emit(f"[{host}] ❌ Excel 数据不完整，跳过")
                continue

            if self.add_cfg["enable"]:
                ssh_create_user(
                    host, user, pwd, port,
                    self.add_cfg["user"],
                    self.add_cfg["pass"],
                    self.add_cfg["sudo"],
                    self.log_signal.emit
                )

            if self.del_cfg["enable"]:
                ssh_delete_user(
                    host, user, pwd, port,
                    self.del_cfg["user"],
                    self.del_cfg["sudo"],
                    self.log_signal.emit
                )

        self.finished_signal.emit()


# ================= GUI =================
class UserManageGUI(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("批量 Linux 用户管理工具")
        self.resize(900, 600)
        self.excel_path = ""
        self._init_logging()
        self.init_ui()

    def _init_logging(self):
        log_dir = os.path.join(os.path.dirname(__file__), "logs")
        os.makedirs(log_dir, exist_ok=True)
        fh = logging.FileHandler(
            os.path.join(log_dir, f"用户管理_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"),
            encoding="utf-8")
        fh.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        logger.addHandler(fh)
        logger.setLevel(logging.INFO)

    def init_ui(self):
        self.setStyleSheet("""
            UserManageGUI { background:#f5f6fa; }
            QGroupBox { font-weight:bold; color:#2d3436; border:none; margin-top:12px; padding:0; }
            QGroupBox::title { padding:0 0 6px 0; border-bottom:2px solid #0984e3; }
            QLineEdit { border:none; border-bottom:1px solid #dfe6e9; padding:6px 4px; background:transparent; font-size:13px; }
            QLineEdit:focus { border-bottom:2px solid #0984e3; }
            QCheckBox { spacing:6px; font-size:13px; }
            QTextEdit { border:1px solid #dfe6e9; border-radius:4px; background:white; font-size:13px; }
        """)
        layout = QVBoxLayout(self)
        layout.setSpacing(8)
        layout.setContentsMargins(28, 24, 28, 24)

        # 文件
        l1 = QHBoxLayout()
        self.excel_label = QLabel("未选择 Excel 文件")
        l1.addWidget(self.excel_label)
        l1.addStretch()
        btn = QPushButton("选择文件")
        btn.setStyleSheet("QPushButton{background:#0984e3;color:white;padding:7px 20px;border:none;border-radius:4px;font-size:13px;}QPushButton:hover{background:#0873c4;}")
        btn.clicked.connect(self.choose_excel)
        l1.addWidget(btn)
        layout.addLayout(l1)

        layout.addSpacing(8)

        # 新增用户
        l2 = QHBoxLayout()
        self.add_check = QCheckBox("新增用户")
        l2.addWidget(self.add_check)
        l2.addWidget(QLabel("用户名"))
        self.add_user = QLineEdit()
        self.add_user.setPlaceholderText("用户名")
        l2.addWidget(self.add_user)
        self.add_pass1 = QLineEdit()
        self.add_pass1.setEchoMode(QLineEdit.Password)
        self.add_pass1.setPlaceholderText("密码")
        l2.addWidget(self.add_pass1)
        self.add_pass2 = QLineEdit()
        self.add_pass2.setEchoMode(QLineEdit.Password)
        self.add_pass2.setPlaceholderText("确认密码")
        l2.addWidget(self.add_pass2)
        self.add_sudo = QCheckBox("sudo")
        l2.addWidget(self.add_sudo)
        l2.addStretch()
        layout.addLayout(l2)

        layout.addSpacing(4)

        # 删除用户
        l3 = QHBoxLayout()
        self.del_check = QCheckBox("删除用户")
        l3.addWidget(self.del_check)
        l3.addWidget(QLabel("用户名"))
        self.del_user = QLineEdit()
        self.del_user.setPlaceholderText("用户名")
        l3.addWidget(self.del_user)
        self.del_sudo = QCheckBox("sudo")
        l3.addWidget(self.del_sudo)
        l3.addStretch()
        layout.addLayout(l3)

        layout.addSpacing(12)

        run_btn = QPushButton("开始执行")
        run_btn.setStyleSheet("QPushButton{background:#27ae60;color:white;font-weight:bold;padding:10px 36px;border:none;border-radius:4px;font-size:14px;}QPushButton:hover{background:#219a52;}QPushButton:disabled{background:#b2bec3;}")
        run_btn.clicked.connect(self.start_task)
        layout.addWidget(run_btn)

        self.log_box = QTextEdit()
        self.log_box.setReadOnly(True)
        self.log_box.setFont(QFont("Consolas", 10))
        layout.addWidget(self.log_box)

    def log(self, msg):
        self.log_box.append(msg)
        self.log_box.ensureCursorVisible()
        logger.info(msg)

    def choose_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择 Excel", "", "Excel (*.xlsx)"
        )
        if path:
            self.excel_path = path
            self.excel_label.setText(os.path.basename(path))

    def start_task(self):
        if not self.excel_path:
            QMessageBox.critical(self, "错误", "请先选择 Excel 文件")
            return

        if not self.add_check.isChecked() and not self.del_check.isChecked():
            QMessageBox.critical(self, "错误", "至少选择一个操作")
            return

        add_cfg = {"enable": False}
        del_cfg = {"enable": False}

        if self.add_check.isChecked():
            if not self.add_user.text():
                QMessageBox.critical(self, "错误", "请输入新增用户名")
                return
            if self.add_pass1.text() != self.add_pass2.text():
                QMessageBox.critical(self, "错误", "密码不一致")
                return
            add_cfg = {
                "enable": True,
                "user": self.add_user.text(),
                "pass": self.add_pass1.text(),
                "sudo": self.add_sudo.isChecked()
            }

        if self.del_check.isChecked():
            if not self.del_user.text():
                QMessageBox.critical(self, "错误", "请输入删除用户名")
                return
            del_cfg = {
                "enable": True,
                "user": self.del_user.text(),
                "sudo": self.del_sudo.isChecked()
            }

        try:
            wb = load_workbook(self.excel_path)
            tasks = list(wb.active.iter_rows(min_row=2, values_only=True))
        except Exception as e:
            QMessageBox.critical(self, "错误", str(e))
            return

        self.thread = QThread()
        self.worker = Worker(tasks, add_cfg, del_cfg)
        self.worker.moveToThread(self.thread)

        self.thread.started.connect(self.worker.run)
        self.worker.log_signal.connect(self.log)
        self.worker.finished_signal.connect(self.thread.quit)
        self.worker.finished_signal.connect(self.worker.deleteLater)
        self.thread.finished.connect(self.thread.deleteLater)
        self.worker.finished_signal.connect(self.on_finished)

        self.thread.start()
        self.log("=== 任务开始执行 ===")

    def on_finished(self):
        self.log("=== 任务执行完成 ===")
        QMessageBox.information(self, "完成", "所有服务器已处理完毕")

# ================= 主入口 =================
if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s",
                        handlers=[logging.StreamHandler()])
    app = QApplication(sys.argv)
    win = UserManageGUI()
    win.show()
    sys.exit(app.exec())
