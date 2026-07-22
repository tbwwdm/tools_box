# -*- coding: utf-8 -*-
"""
Linux 工具集 — 统一 Tab 界面
布局：功能区 → Excel导入+服务器表格 → 执行按钮 → 日志
"""
import sys, os, re, logging, socket, traceback, threading
from datetime import datetime

import paramiko
import pandas as pd
from openpyxl import load_workbook, Workbook
from concurrent.futures import ThreadPoolExecutor, as_completed

from PySide6.QtWidgets import (
    QApplication, QWidget, QPushButton, QLabel, QLineEdit,
    QTextEdit, QFileDialog, QVBoxLayout, QHBoxLayout,
    QMessageBox, QCheckBox, QComboBox, QTabWidget,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QDialog, QFormLayout, QStyle, QStyleOptionButton, QMenu,
    QSplitter
)
from PySide6.QtCore import QObject, Signal, QThread, Qt, QRect, QTimer
from PySide6.QtGui import QFont

logger = logging.getLogger(__name__)
THIS_DIR = os.path.dirname(__file__)


class _GuiLogHandler(logging.Handler):
    """将 logging 输出转发到 Qt 信号，用于界面日志框显示（过滤掉 paramiko 等干扰日志）"""
    _SKIP_PREFIXES = ("paramiko.transport",)

    def __init__(self, signal):
        super().__init__()
        self._signal = signal

    def emit(self, record):
        if record.name.startswith(self._SKIP_PREFIXES):
            return
        try:
            msg = self.format(record)
            self._signal.emit(msg)
        except Exception:
            self.handleError(record)


# ================= 共享 Excel 读取 =================
def read_excel_tasks(filepath):
    wb = load_workbook(filepath)
    tasks = []
    for row in wb.active.iter_rows(min_row=2, values_only=True):
        host = str(row[0]) if row[0] is not None else ""
        port = int(row[1]) if len(row) > 1 and row[1] is not None else 22
        user = str(row[2]) if len(row) > 2 and row[2] is not None else ""
        pwd = str(row[3]) if len(row) > 3 and row[3] is not None else ""
        if host:
            tasks.append((host, port, user, pwd))
    return tasks


# ================= SSH 操作 =================
def ssh_create_user(ip, login_user, login_pass, port, new_user, new_pass, use_sudo):
    logger.info(f"[{ip}] 连接中: {login_user}@{ip}:{port}")
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(ip, int(port), login_user, login_pass,
                timeout=10, look_for_keys=False, allow_agent=False)
    logger.info(f"[{ip}] 连接成功，检查用户 {new_user} 是否存在")
    prefix = "sudo " if use_sudo else ""
    stdin, stdout, _ = ssh.exec_command(f"id {new_user}")
    if stdout.channel.recv_exit_status() == 0:
        ssh.close()
        logger.info(f"[{ip}] 用户 {new_user} 已存在，跳过")
        return f"[{ip}] 用户 {new_user} 已存在，跳过"
    logger.info(f"[{ip}] 创建用户 {new_user}")
    ssh.exec_command(f"{prefix}useradd {new_user}")
    ssh.exec_command(f'echo "{new_user}:{new_pass}" | {prefix}chpasswd')
    ssh.close()
    logger.info(f"[{ip}] ✅ 用户 {new_user} 创建成功")
    return f"[{ip}] ✅ 用户 {new_user} 创建成功"


def ssh_delete_user(ip, login_user, login_pass, port, del_user, use_sudo):
    logger.info(f"[{ip}] 连接中: {login_user}@{ip}:{port}")
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(ip, int(port), login_user, login_pass,
                timeout=10, look_for_keys=False, allow_agent=False)
    logger.info(f"[{ip}] 连接成功，检查用户 {del_user} 是否存在")
    prefix = "sudo " if use_sudo else ""
    stdin, stdout, _ = ssh.exec_command(f"id {del_user}")
    if stdout.channel.recv_exit_status() != 0:
        ssh.close()
        logger.info(f"[{ip}] 用户 {del_user} 不存在，跳过")
        return f"[{ip}] 用户 {del_user} 不存在，跳过"
    logger.info(f"[{ip}] 删除用户 {del_user}")
    ssh.exec_command(f"{prefix}userdel -r {del_user}")
    ssh.close()
    logger.info(f"[{ip}] ✅ 用户 {del_user} 已删除")
    return f"[{ip}] ✅ 用户 {del_user} 已删除"


def ssh_modify_sshd_config(ip, login_user, login_pass, port, sshd_options):
    logger.info(f"[{ip}] 连接中: {login_user}@{ip}:{port}")
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(hostname=ip, port=int(port), username=login_user,
                password=login_pass, timeout=10,
                look_for_keys=False, allow_agent=False)
    logger.info(f"[{ip}] 连接成功，执行 sshd 配置修改")
    cmds = ["cp /etc/ssh/sshd_config /etc/ssh/sshd_config.bak"]
    for key, value in sshd_options.items():
        cmds.append(f"sed -i '/^#\\?{key} /d' /etc/ssh/sshd_config")
        cmds.append(f"echo '{key} {value}' >> /etc/ssh/sshd_config")
    cmds.append("systemctl restart sshd")
    for cmd in cmds:
        logger.info(f"[{ip}] 执行: {cmd}")
        stdin, stdout, stderr = ssh.exec_command(cmd)
        err = stderr.read().decode().strip()
        if err:
            logger.error(f"[{ip}] 命令失败: {cmd} → {err}")
            ssh.close()
            return f"[{ip}] ❌ 失败: {err}"
    ssh.close()
    changes = []
    for k, v in sshd_options.items():
        if k == "Port":
            changes.append(f"端口修改为 {v}")
        elif k == "PermitRootLogin":
            changes.append("禁止 root 登录" if v.lower() == "no" else "允许 root 登录")
    logger.info(f"[{ip}] ✅ sshd 配置成功: {', '.join(changes)}")
    return f"[{ip}] ✅ sshd 配置成功，" + "；".join(changes)


def ssh_test_login(ip, login_user, login_pass, port):
    logger.info(f"[{ip}] 测试登录: {login_user}@{ip}:{port}")
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(ip, port=int(port), username=login_user, password=login_pass,
                timeout=3, look_for_keys=False, allow_agent=False)
    ssh.close()
    logger.info(f"[{ip}] ✅ 登录成功")
    return f"[{ip}] ✅ 登录成功"


def ssh_get_userlist(ip, login_user, login_pass, port):
    logger.info(f"[{ip}] 连接中: {login_user}@{ip}:{port}")
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(ip, port=int(port), username=login_user, password=login_pass,
                timeout=10, look_for_keys=False, allow_agent=False)
    logger.info(f"[{ip}] 连接成功，获取用户列表")
    _, stdout, _ = ssh.exec_command(
        """awk -F: '$3>=1000 && $7!="/sbin/nologin" && $7!="/usr/sbin/nologin" {print $1}' /etc/passwd""")
    users = stdout.read().decode().splitlines()
    ssh.close()
    logger.info(f"[{ip}] 获取到 {len(users)} 个用户: {', '.join(users)}")
    return users


# ================= 安全基线 =================
def _ssh_connect(host, username, password):
    logger.info(f"[{host}] 连接安全扫描: {username}@{host}")
    c = paramiko.SSHClient()
    c.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    c.connect(host, username=username, password=password,
              timeout=10, look_for_keys=False, allow_agent=False)
    logger.info(f"[{host}] 连接成功")
    return c


def _check_os_type(ssh):
    o = ssh.exec_command('cat /etc/os-release')[1].read().decode().lower()
    if 'ubuntu' in o:
        logger.info(f"检测到系统类型: Ubuntu")
        return 'ubuntu'
    if 'centos' in o or 'red hat' in o:
        logger.info(f"检测到系统类型: RedHat/CentOS")
        return 'redhat'
    if 'kylin' in o:
        logger.info(f"检测到系统类型: Kylin")
        return 'kylin'
    logger.warning(f"无法识别的系统类型 ({o[:50]}...)，按 RedHat 处理")
    return 'other'


def _add_result(results, ip, cat, item, ok, detail):
    results.append((cat, item, ok, detail))


def _check_redhat(ssh, ip):
    r = []
    lo = ssh.exec_command('cat /etc/login.defs')[1].read().decode()
    m = re.search(r'^\s*PASS_MIN_LEN\s+(\d+)', lo, re.M | re.I)
    _add_result(r, ip, '密码策略', '密码最小长度',
                '通过' if m and int(m.group(1)) >= 8 else '未通过',
                f'PASS_MIN_LEN={m.group(1)}' if m else '需设置PASS_MIN_LEN>=8')
    m = re.search(r'^\s*PASS_MAX_DAYS\s+(\d+)', lo, re.M | re.I)
    _add_result(r, ip, '密码策略', '密码过期时间',
                '通过' if m and int(m.group(1)) <= 90 else '未通过',
                f'PASS_MAX_DAYS={m.group(1)}' if m else '需设置PASS_MAX_DAYS<=90')
    pa = ssh.exec_command('cat /etc/pam.d/system-auth')[1].read().decode()
    ok = all(x in pa for x in ['retry=3', 'minlen=8', 'minclass=3'])
    _add_result(r, ip, '密码策略', '密码创建要求', '通过' if ok else '未通过',
                '配置正确' if ok else '需配置pam_cracklib')
    rs = ssh.exec_command('cat /etc/rsyslog.conf')[1].read().decode()
    _add_result(r, ip, '日志配置', '日志服务器',
                '通过' if re.search(r'\*\.\*\s+@\d+', rs) else '未通过',
                '已配置' if re.search(r'\*\.\*\s+@\d+', rs) else '需配置*.*@IP')
    _add_result(r, ip, '日志配置', 'cron日志',
                '通过' if 'cron.* /var/log/cron' in rs else '未通过',
                '已记录' if 'cron.* /var/log/cron' in rs else '需配置cron.* /var/log/cron')
    _add_result(r, ip, '日志配置', 'Syslog审计',
                '通过' if 'authpriv.* /var/log/secure' in rs else '未通过',
                '已启用' if 'authpriv.* /var/log/secure' in rs else '需配置authpriv.* /var/log/secure')
    pf = ssh.exec_command('cat /etc/profile')[1].read().decode()
    _add_result(r, ip, '登录超时', '超时设置',
                '通过' if 'TMOUT=600' in pf else '未通过',
                'TMOUT=600已配置' if 'TMOUT=600' in pf else '需在/etc/profile设置TMOUT=600')
    for f, p in [('/etc/passwd', '644'), ('/etc/group', '644'), ('/etc/shadow', '600')]:
        a = ssh.exec_command(f'stat -c "%a" {f}')[1].read().decode().strip()
        _add_result(r, ip, '权限控制', f'{f}权限',
                    '通过' if a == p else '未通过',
                    f'权限{a}' if a == p else f'需设置权限{p}')
    _add_result(r, ip, '权限控制', '缺省umask',
                '通过' if re.search(r'^\s*umask\s+027\s*$', lo, re.M) else '未通过',
                'umask=027' if re.search(r'^\s*umask\s+027\s*$', lo, re.M) else '需设置umask 027')
    v = ssh.exec_command('systemctl is-active vsftpd')[1].read().decode().strip()
    if v == 'active':
        ft = ssh.exec_command('cat /etc/vsftpd/ftpusers')[1].read().decode()
        _add_result(r, ip, 'FTP设置', '禁止root登录FTP',
                    '通过' if 'root' in ft else '未通过',
                    '已禁止' if 'root' in ft else '需在ftpusers中禁止root')
        cf = ssh.exec_command('cat /etc/vsftpd/vsftpd.conf')[1].read().decode()
        _add_result(r, ip, 'FTP设置', '禁止匿名FTP',
                    '通过' if re.search(r'anonymous_enable\s*=\s*no', cf, re.I) else '未通过',
                    '已禁止' if re.search(r'anonymous_enable\s*=\s*no', cf, re.I) else '需设置anonymous_enable=no')
    else:
        _add_result(r, ip, 'FTP设置', '禁止root登录FTP', '通过', 'VSFTPD未运行')
        _add_result(r, ip, 'FTP设置', '禁止匿名FTP', '通过', 'VSFTPD未运行')
    t = ssh.exec_command('systemctl is-active xinetd.service')[1].read().decode().strip()
    _add_result(r, ip, '服务检查', '禁用Telnet',
                '通过' if t != 'active' else '未通过',
                '未启用' if t != 'active' else '建议关闭telnet')
    return r


def _check_ubuntu(ssh, ip):
    r = []
    lo = ssh.exec_command('cat /etc/login.defs')[1].read().decode()
    m = re.search(r'^\s*PASS_MIN_LEN\s+(\d+)', lo, re.M | re.I)
    _add_result(r, ip, '密码策略', '密码最小长度',
                '通过' if m and int(m.group(1)) >= 8 else '未通过',
                f'PASS_MIN_LEN={m.group(1)}' if m else '需设置PASS_MIN_LEN>=8')
    m = re.search(r'^\s*PASS_MAX_DAYS\s+(\d+)', lo, re.M | re.I)
    _add_result(r, ip, '密码策略', '密码过期时间',
                '通过' if m and int(m.group(1)) <= 90 else '未通过',
                f'PASS_MAX_DAYS={m.group(1)}' if m else '需设置PASS_MAX_DAYS<=90')
    pa = ssh.exec_command('cat /etc/pam.d/common-password')[1].read().decode()
    ok = all(x in pa for x in ['retry=3', 'minlen=8', 'minclass=3'])
    _add_result(r, ip, '密码策略', '密码创建要求', '通过' if ok else '未通过',
                '配置正确' if ok else '需配置pam_cracklib')
    rs = ssh.exec_command('cat /etc/rsyslog.conf')[1].read().decode()
    _add_result(r, ip, '日志配置', '日志服务器',
                '通过' if re.search(r'\*\.\*\s+@\d+', rs) else '未通过',
                '已配置' if re.search(r'\*\.\*\s+@\d+', rs) else '需配置*.*@IP')
    _add_result(r, ip, '日志配置', 'cron日志',
                '通过' if 'cron.* /var/log/cron' in rs else '未通过',
                '已记录' if 'cron.* /var/log/cron' in rs else '需配置cron.* /var/log/cron')
    _add_result(r, ip, '日志配置', 'Syslog审计',
                '通过' if 'authpriv.* /var/log/secure' in rs else '未通过',
                '已启用' if 'authpriv.* /var/log/secure' in rs else '需配置authpriv.* /var/log/secure')
    pf = ssh.exec_command('cat /etc/profile')[1].read().decode()
    _add_result(r, ip, '登录超时', '超时设置',
                '通过' if 'TMOUT=600' in pf else '未通过',
                'TMOUT=600已配置' if 'TMOUT=600' in pf else '需在/etc/profile设置TMOUT=600')
    for f, p in [('/etc/passwd', '644'), ('/etc/group', '644'), ('/etc/shadow', '600')]:
        a = ssh.exec_command(f'stat -c "%a" {f}')[1].read().decode().strip()
        _add_result(r, ip, '权限控制', f'{f}权限',
                    '通过' if a == p else '未通过',
                    f'权限{a}' if a == p else f'需设置权限{p}')
    _add_result(r, ip, '权限控制', '缺省umask',
                '通过' if re.search(r'^\s*umask\s+027\s*$', lo, re.M) else '未通过',
                'umask=027' if re.search(r'^\s*umask\s+027\s*$', lo, re.M) else '需设置umask 027')
    v = ssh.exec_command('systemctl is-active vsftpd')[1].read().decode().strip()
    if v == 'active':
        ft = ssh.exec_command('cat /etc/vsftpd/ftpusers')[1].read().decode()
        _add_result(r, ip, 'FTP设置', '禁止root登录FTP',
                    '通过' if 'root' in ft else '未通过',
                    '已禁止' if 'root' in ft else '需在ftpusers中禁止root')
        cf = ssh.exec_command('cat /etc/vsftpd/vsftpd.conf')[1].read().decode()
        _add_result(r, ip, 'FTP设置', '禁止匿名FTP',
                    '通过' if re.search(r'anonymous_enable\s*=\s*no', cf, re.I) else '未通过',
                    '已禁止' if re.search(r'anonymous_enable\s*=\s*no', cf, re.I) else '需设置anonymous_enable=no')
    else:
        _add_result(r, ip, 'FTP设置', '禁止root登录FTP', '通过', 'VSFTPD未运行')
        _add_result(r, ip, 'FTP设置', '禁止匿名FTP', '通过', 'VSFTPD未运行')
    t = ssh.exec_command('systemctl is-active xinetd.service')[1].read().decode().strip()
    _add_result(r, ip, '服务检查', '禁用Telnet',
                '通过' if t != 'active' else '未通过',
                '未启用' if t != 'active' else '建议关闭telnet')
    return r


def _check_host(ssh, os_type, ip):
    if os_type in ('redhat', 'kylin'):
        return _check_redhat(ssh, ip)
    elif os_type == 'ubuntu':
        return _check_ubuntu(ssh, ip)
    return [('系统检查', '操作系统类型', '未通过', '不支持')]


def _save_security_results(excel_path, host, ip, results, st, et):
    df = pd.DataFrame([{
        '主机名称': host, 'IP': ip, '检查大类': r[0], '检查项': r[1],
        '检查结果': r[2], '加固建议': r[3], '开始时间': st, '结束时间': et
    } for r in results])
    if os.path.exists(excel_path):
        df = pd.concat([pd.read_excel(excel_path), df], ignore_index=True)
    df.to_excel(excel_path, index=False)


def _scan_one(row_data, excel_path, log_cb, cancel_flag):
    host, port, user, pwd = row_data
    if cancel_flag[0]:
        return
    st = datetime.now()
    log_cb(f"[{host}] ─── 开始检查 ───")
    try:
        ssh = _ssh_connect(host, user, pwd)
        ot = _check_os_type(ssh)
        res = _check_host(ssh, ot, host)
        _save_security_results(excel_path, host, host, res, st, datetime.now())
        ssh.close()
        pass_cnt = sum(1 for r in res if r[2] == '通过')
        for cat, item, ok, detail in res:
            icon = "✅" if ok == "通过" else "❌"
            log_cb(f"[{host}]  [{cat}] {icon} {item}  →  {detail}")
        log_cb(f"[{host}] ✅ 完成 ({pass_cnt}/{len(res)} 项通过)")
    except Exception as e:
        tb = traceback.format_exc()
        log_cb(f"[{host}] ❌ 失败: {e}\n{tb}")


# ================= Workers（支持取消） =================
class _CancelFlag:
    def __init__(self):
        self.cancelled = False


class _UserWorker(QObject):
    log_signal = Signal(str)
    finished_signal = Signal()

    def __init__(self, tasks, add_cfg, del_cfg, cancel_flag):
        super().__init__()
        self.tasks = tasks
        self.add_cfg = add_cfg
        self.del_cfg = del_cfg
        self._cancel = cancel_flag

    def run(self):
        try:
            for host, port, user, pwd in self.tasks:
                if self._cancel.cancelled:
                    self.log_signal.emit("任务已取消")
                    return
                if not all([host, user, pwd]):
                    self.log_signal.emit(f"[{host}] ❌ 数据不完整，跳过")
                    continue
                try:
                    if self.add_cfg.get("enable"):
                        msg = ssh_create_user(host, user, pwd, port,
                                              self.add_cfg["user"],
                                              self.add_cfg["pass"],
                                              self.add_cfg["sudo"])
                        self.log_signal.emit(msg)
                    if self.del_cfg.get("enable"):
                        msg = ssh_delete_user(host, user, pwd, port,
                                              self.del_cfg["user"],
                                              self.del_cfg["sudo"])
                        self.log_signal.emit(msg)
                except (socket.error, Exception) as e:
                    tb = traceback.format_exc()
                    self.log_signal.emit(f"[{host}] ❌ 连接失败: {e}\n{tb}")
        finally:
            self.finished_signal.emit()


class _SSHConfigWorker(QObject):
    log_signal = Signal(str)
    finished_signal = Signal()

    def __init__(self, tasks, sshd_options, cancel_flag):
        super().__init__()
        self.tasks = tasks
        self.sshd_options = sshd_options
        self._cancel = cancel_flag

    def run(self):
        try:
            for host, port, user, pwd in self.tasks:
                if self._cancel.cancelled:
                    self.log_signal.emit("任务已取消")
                    return
                if not all([host, user, pwd]):
                    self.log_signal.emit(f"[{host}] ❌ 数据不完整，跳过")
                    continue
                try:
                    msg = ssh_modify_sshd_config(host, user, pwd, port, self.sshd_options)
                    self.log_signal.emit(msg)
                except (socket.error, Exception) as e:
                    tb = traceback.format_exc()
                    self.log_signal.emit(f"[{host}] ❌ 失败: {e}\n{tb}")
        finally:
            self.finished_signal.emit()


class _UserListWorker(QObject):
    log_signal = Signal(str)
    finished_signal = Signal()

    def __init__(self, tasks, output_path, cancel_flag):
        super().__init__()
        self.tasks = tasks
        self.output_path = output_path
        self._cancel = cancel_flag

    def run(self):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "用户列表"
            # 表头
            ws.cell(row=1, column=1, value="Host")
            ws.cell(row=1, column=2, value="Port")
            for i, (host, port, user, pwd) in enumerate(self.tasks):
                if self._cancel.cancelled:
                    self.log_signal.emit("任务已取消")
                    return
                if not all([host, user, pwd]):
                    self.log_signal.emit(f"[{host}] ❌ 数据不完整，跳过")
                    continue
                try:
                    users = ssh_get_userlist(host, user, pwd, port)
                    row_excel = i + 2
                    ws.cell(row=row_excel, column=1, value=host)
                    ws.cell(row=row_excel, column=2, value=port)
                    for j, u in enumerate(users[:20]):
                        # 每个用户占一列，从第3列开始
                        col = 3 + j
                        if row_excel == 2:  # 第一行数据时写表头
                            ws.cell(row=1, column=col, value=f"用户{j+1}")
                        ws.cell(row=row_excel, column=col, value=u)
                    self.log_signal.emit(f"[{host}] ✅ 获取到 {len(users)} 个用户: {', '.join(users)}")
                except Exception as e:
                    tb = traceback.format_exc()
                    self.log_signal.emit(f"[{host}] ❌ 失败: {e}\n{tb}")
            wb.save(self.output_path)
            self.log_signal.emit(f"完成！结果已写入 {self.output_path}")
        except Exception as e:
            tb = traceback.format_exc()
            self.log_signal.emit(f"错误: {e}\n{tb}")
        finally:
            self.finished_signal.emit()


class _SecurityWorker(QObject):
    log_signal = Signal(str)
    finished_signal = Signal()

    def __init__(self, tasks, excel_path, max_workers, cancel_flag):
        super().__init__()
        self.tasks = tasks
        self.excel_path = excel_path
        self.max_workers = max_workers
        self._cancel = cancel_flag

    def run(self):
        try:
            cancelled = [False]
            self.log_signal.emit(f"开始扫描 {len(self.tasks)} 台服务器，并发 {self.max_workers}")

            def check_cancel():
                if self._cancel.cancelled:
                    cancelled[0] = True

            with ThreadPoolExecutor(max_workers=self.max_workers) as ex:
                futures = {}
                for t in self.tasks:
                    if self._cancel.cancelled:
                        break
                    futures[ex.submit(
                        _scan_one, t, self.excel_path,
                        self.log_signal.emit, cancelled
                    )] = t
                for f in as_completed(futures):
                    if self._cancel.cancelled:
                        ex.shutdown(wait=False, cancel_futures=True)
                        self.log_signal.emit("⚠️ 扫描已中断")
                        break
            if not self._cancel.cancelled:
                self.log_signal.emit("✅ 扫描完成")
        except Exception as e:
            tb = traceback.format_exc()
            self.log_signal.emit(f"扫描失败: {e}\n{tb}")
        finally:
            self.finished_signal.emit()


# ================= 主界面 =================
COL_CHECK = 0
COL_HOST = 1
COL_PORT = 2
COL_USER = 3
COL_PWD = 4
COL_STATUS = 5
HEADERS = ["", "Host", "端口", "用户名", "密码", "状态"]


class _ManualAddDialog(QDialog):
    """手动添加服务器的弹出对话框"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("手动添加服务器")
        self.setMinimumWidth(400)
        self._result = None
        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        fl = QFormLayout()
        fl.setSpacing(8)
        self._host_edit = QLineEdit()
        self._host_edit.setPlaceholderText("IP 地址或主机名")
        fl.addRow("Host", self._host_edit)

        self._port_edit = QLineEdit("22")
        self._port_edit.setPlaceholderText("SSH 端口")
        fl.addRow("端口", self._port_edit)

        self._user_edit = QLineEdit()
        self._user_edit.setPlaceholderText("登录用户名")
        fl.addRow("用户名", self._user_edit)

        self._pwd_edit = QLineEdit()
        self._pwd_edit.setEchoMode(QLineEdit.Password)
        self._pwd_edit.setPlaceholderText("登录密码")
        fl.addRow("密码", self._pwd_edit)
        layout.addLayout(fl)

        # 按钮行
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        ok_btn = QPushButton("确定")
        ok_btn.setStyleSheet(
            "QPushButton{background:#0984e3;color:white;padding:7px 28px;"
            "border:none;border-radius:4px;font-size:13px;}"
            "QPushButton:hover{background:#0873c4;}")
        ok_btn.clicked.connect(self._on_ok)
        btn_row.addWidget(ok_btn)
        cancel_btn = QPushButton("取消")
        cancel_btn.setStyleSheet(
            "QPushButton{background:#636e72;color:white;padding:7px 28px;"
            "border:none;border-radius:4px;font-size:13px;}"
            "QPushButton:hover{background:#50575a;}")
        cancel_btn.clicked.connect(self.reject)
        btn_row.addWidget(cancel_btn)
        layout.addLayout(btn_row)

        self._host_edit.setFocus()

    def _on_ok(self):
        host = self._host_edit.text().strip()
        port_str = self._port_edit.text().strip()
        user = self._user_edit.text().strip()
        pwd = self._pwd_edit.text()

        if not host:
            QMessageBox.warning(self, "提示", "请输入 Host")
            self._host_edit.setFocus()
            return
        if not user:
            QMessageBox.warning(self, "提示", "请输入用户名")
            self._user_edit.setFocus()
            return
        if not pwd:
            QMessageBox.warning(self, "提示", "请输入密码")
            self._pwd_edit.setFocus()
            return
        try:
            port = int(port_str) if port_str else 22
            if not (1 <= port <= 65535):
                raise ValueError
        except ValueError:
            QMessageBox.warning(self, "提示", "端口必须是 1-65535 之间的数字")
            self._port_edit.setFocus()
            return

        self._result = (host, port, user, pwd)
        self.accept()

    def get_result(self):
        return self._result


class _SelectAllHeader(QHeaderView):
    """表头第一列绘制原生复选框，与表格内复选框风格一致"""
    def __init__(self, parent=None):
        super().__init__(Qt.Horizontal, parent)
        self._checked = False
        self.setSectionsClickable(True)

    def paintSection(self, painter, rect, logicalIndex):
        painter.save()
        super().paintSection(painter, rect, logicalIndex)
        if logicalIndex == 0:
            btn_opt = QStyleOptionButton()
            sz = self.style().pixelMetric(QStyle.PM_IndicatorWidth, None, self)
            if sz <= 0:
                sz = 16
            x = rect.x() + (rect.width() - sz) // 2
            y = rect.y() + (rect.height() - sz) // 2
            btn_opt.rect = QRect(int(x), int(y), sz, sz)
            btn_opt.state = QStyle.State_Enabled
            btn_opt.state |= QStyle.State_On if self._checked else QStyle.State_Off
            self.style().drawControl(QStyle.CE_CheckBox, btn_opt, painter)
        painter.restore()

    def set_checked(self, checked):
        if self._checked != checked:
            self._checked = checked
            self.updateSection(0)

    def is_checked(self):
        return self._checked


class LinuxToolsGUI(QWidget):
    _gui_log_signal = Signal(str)

    def __init__(self, lang="zh"):
        super().__init__()
        self.lang = lang
        title = "Linux Tools Suite" if self.lang == "en" else "Linux 工具集"
        self.setWindowTitle(title)
        self.resize(1000, 700)
        self._excel_path = ""
        self._tab_pages = []
        self._thread = None
        self._worker = None
        self._cancel_flag = None
        self._init_logging()
        self._init_ui()
        self._gui_log_signal.connect(self._auto_log)
        self._status_signal.connect(self._on_status_update)

    def _init_logging(self):
        log_dir = os.path.join(THIS_DIR, "logs")
        os.makedirs(log_dir, exist_ok=True)
        fh = logging.FileHandler(
            os.path.join(log_dir, f"Linux工具集_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"),
            encoding="utf-8")
        fh.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        logger.addHandler(fh)
        logger.setLevel(logging.INFO)
        # 阻止传播到 root，避免 GUI 消息重复（GUI 已由 _log/_auto_log 直接输出）
        logger.propagate = False

        # 将 root logger 的所有输出（含 paramiko 等第三方库）转发到界面日志框
        gui_handler = _GuiLogHandler(self._gui_log_signal)
        gui_handler.setFormatter(logging.Formatter('%(name)s [%(levelname)s] %(message)s'))
        gui_handler.setLevel(logging.INFO)
        logging.getLogger().addHandler(gui_handler)

    def _init_ui(self):
        self.setStyleSheet("""
            LinuxToolsGUI { background:#f5f6fa; }
            QTabWidget::pane { border:1px solid #dfe6e9; border-radius:6px; background:white; }
            QTabBar::tab { background:#ecf0f1; color:#636e72; padding:10px 22px;
                           border:1px solid #dfe6e9; border-bottom:none;
                           border-top-left-radius:6px; border-top-right-radius:6px;
                           font-size:13px; margin-right:2px; }
            QTabBar::tab:selected { background:white; color:#0984e3; font-weight:bold; }
            QTabBar::tab:hover { background:#dfe6e9; }
            QLineEdit { border:none; border-bottom:1px solid #dfe6e9; padding:6px 4px;
                        background:transparent; font-size:13px; }
            QLineEdit:focus { border-bottom:2px solid #0984e3; }
            QCheckBox { spacing:6px; font-size:13px; }
            QComboBox { border:none; border-bottom:1px solid #dfe6e9; padding:6px 4px;
                        background:transparent; font-size:13px; }
            QComboBox:focus { border-bottom:2px solid #0984e3; }
            QTextEdit { border:1px solid #dfe6e9; border-radius:4px; background:white; font-size:13px; }
            QTableWidget { border:1px solid #dfe6e9; border-radius:4px; background:white;
                           gridline-color:#f0f0f0; font-size:13px; }
            QTableWidget::item { padding:4px 8px; }
            QTableWidget::item:selected { background:#e8f0fe; color:#2d3436; }
            QHeaderView::section { background:#f5f6fa; color:#636e72;                                    padding:8px 4px; border:none; border-bottom:2px solid #dfe6e9; }
        """)
        layout = QVBoxLayout(self)
        layout.setSpacing(6)
        layout.setContentsMargins(20, 16, 20, 16)

        # ===== 1. 功能区（Tab，内含独立导入栏+表格） =====
        self._tabs = QTabWidget()
        self._tabs.addTab(self._create_ssh_tab(), "SSH配置修改")
        self._tabs.addTab(self._create_user_tab(), "用户新增删除")
        self._tabs.addTab(self._create_userlist_tab(), "获取用户列表")
        self._tabs.addTab(self._create_security_tab(), "安全基线扫描")
        self._tabs.currentChanged.connect(self._on_tab_changed)
        # 初始化第一个 Tab 的表格引用
        if self._tab_pages:
            self._on_tab_changed(0)

        # ===== 4. 执行 / 取消 =====
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        self._run_btn = QPushButton("▶ Run" if self.lang == "en" else "▶ 执行")
        self._run_btn.setStyleSheet(
            "QPushButton{background:#27ae60;color:white;padding:7px 18px;"
            "border:none;border-radius:4px;font-size:13px;}"
            "QPushButton:hover{background:#219a52;}"
            "QPushButton:disabled{background:#b2bec3;}")
        self._run_btn.clicked.connect(self._start)
        btn_row.addWidget(self._run_btn)

        self._cancel_btn = QPushButton("✕ Cancel" if self.lang == "en" else "✕ 取消")
        self._cancel_btn.setEnabled(False)
        self._cancel_btn.setStyleSheet(
            "QPushButton{background:#e74c3c;color:white;padding:7px 18px;"
            "border:none;border-radius:4px;font-size:13px;}"
            "QPushButton:hover{background:#c0392b;}"
            "QPushButton:disabled{background:#b2bec3;}")
        self._cancel_btn.clicked.connect(self._cancel)
        btn_row.addWidget(self._cancel_btn)

        self._copy_btn = QPushButton("Copy Log" if self.lang == "en" else "复制日志")
        self._copy_btn.setStyleSheet(
            "QPushButton{background:#636e72;color:white;padding:7px 18px;"
            "border:none;border-radius:4px;font-size:13px;}"
            "QPushButton:hover{background:#50575a;}")
        self._copy_btn.clicked.connect(self._copy_log)
        btn_row.addWidget(self._copy_btn)

        self._clear_btn = QPushButton("Clear Log" if self.lang == "en" else "清空日志")
        self._clear_btn.setStyleSheet(
            "QPushButton{background:#e74c3c;color:white;padding:7px 18px;"
            "border:none;border-radius:4px;font-size:13px;}"
            "QPushButton:hover{background:#c0392b;}")
        self._clear_btn.clicked.connect(self._clear_log)
        btn_row.addWidget(self._clear_btn)

        btn_row.addStretch()

        # ===== 5. 将功能区和按钮放入容器，与日志框可拖拽分隔 =====
        top_widget = QWidget()
        top_layout = QVBoxLayout(top_widget)
        top_layout.setContentsMargins(0, 0, 0, 0)
        top_layout.addWidget(self._tabs, 1)
        top_layout.addLayout(btn_row)

        self._log_box = QTextEdit()
        self._log_box.setReadOnly(True)
        self._log_box.setAcceptRichText(True)
        self._log_box.setFont(QFont("Consolas", 10))

        self._splitter = QSplitter(Qt.Vertical)
        self._splitter.addWidget(top_widget)
        self._splitter.addWidget(self._log_box)
        self._splitter.setStretchFactor(0, 2)
        self._splitter.setStretchFactor(1, 1)
        self._splitter.setChildrenCollapsible(False)
        layout.addWidget(self._splitter, 1)

    # ===================== 共享组件 =====================

    def _create_server_table(self):
        table = QTableWidget(0, 6)
        # 自定义表头（第一列绘制原生复选框）
        header = _SelectAllHeader(table)
        table.setHorizontalHeader(header)
        for i, text in enumerate(HEADERS):
            table.setHorizontalHeaderItem(i, QTableWidgetItem(text))
        table.setColumnHidden(COL_PWD, True)
        header.sectionClicked.connect(
            lambda idx, t=table: self._on_header_check_clicked(idx, t))
        hh = table.horizontalHeader()
        hh.setSectionResizeMode(COL_CHECK, QHeaderView.Fixed)
        hh.setSectionResizeMode(COL_HOST, QHeaderView.Stretch)
        hh.setSectionResizeMode(COL_PORT, QHeaderView.Stretch)
        hh.setSectionResizeMode(COL_USER, QHeaderView.Stretch)
        hh.setSectionResizeMode(COL_STATUS, QHeaderView.Fixed)
        table.setColumnWidth(COL_CHECK, 36)
        table.setColumnWidth(COL_STATUS, 152)
        # 端口、用户名、状态列表头居中显示
        for col in (COL_PORT, COL_USER, COL_STATUS):
            item = table.horizontalHeaderItem(col)
            if item:
                item.setTextAlignment(Qt.AlignCenter)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setSelectionMode(QAbstractItemView.SingleSelection)
        table.verticalHeader().setVisible(False)
        # 右键菜单
        table.setContextMenuPolicy(Qt.CustomContextMenu)
        table.customContextMenuRequested.connect(
            lambda pos, t=table: self._show_table_context_menu(pos, t))
        return table

    def _create_import_bar(self):
        bar = QHBoxLayout()
        bar.setSpacing(10)
        bar.addWidget(QLabel("Excel 文件"))
        excel_label = QLabel("未选择")
        excel_label.setStyleSheet("color:#636e72;font-size:12px;")
        bar.addWidget(excel_label, 1)

        btn_browse = QPushButton("浏览导入")
        btn_browse.setStyleSheet(
            "QPushButton{background:#0984e3;color:white;padding:7px 18px;"
            "border:none;border-radius:4px;font-size:13px;}"
            "QPushButton:hover{background:#0873c4;}")
        btn_browse.clicked.connect(self._import_excel)
        bar.addWidget(btn_browse)

        btn_add = QPushButton("+ 手动添加")
        btn_add.setStyleSheet(
            "QPushButton{background:#636e72;color:white;padding:7px 18px;"
            "border:none;border-radius:4px;font-size:13px;}"
            "QPushButton:hover{background:#50575a;}")
        btn_add.clicked.connect(self._add_manual_row)
        bar.addWidget(btn_add)

        btn_check = QPushButton("状态检测")
        btn_check.setStyleSheet(
            "QPushButton{background:#f39c12;color:white;padding:7px 14px;"
            "border:none;border-radius:4px;font-size:12px;}"
            "QPushButton:hover{background:#d68910;}")
        btn_check.clicked.connect(self._check_ssh_status)
        bar.addWidget(btn_check)

        btn_del = QPushButton("删除勾选")
        btn_del.setStyleSheet(
            "QPushButton{background:#e74c3c;color:white;padding:7px 14px;"
            "border:none;border-radius:4px;font-size:12px;}"
            "QPushButton:hover{background:#c0392b;}")
        btn_del.clicked.connect(self._delete_checked)
        bar.addWidget(btn_del)

        return bar, excel_label

    # ===================== 各 Tab 配置 =====================

    def _create_user_tab(self):
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(4)
        lay.setContentsMargins(20, 10, 20, 8)

        self._add_check = QCheckBox("新增用户")
        self._add_user = QLineEdit(); self._add_user.setPlaceholderText("用户名")
        self._add_pass1 = QLineEdit(); self._add_pass1.setEchoMode(QLineEdit.Password)
        self._add_pass1.setPlaceholderText("密码")
        self._add_pass2 = QLineEdit(); self._add_pass2.setEchoMode(QLineEdit.Password)
        self._add_pass2.setPlaceholderText("确认密码")
        self._add_sudo = QCheckBox("sudo")

        r1 = QHBoxLayout()
        r1.addWidget(self._add_check)
        r1.addWidget(QLabel("用户名")); r1.addWidget(self._add_user)
        r1.addWidget(QLabel("密码")); r1.addWidget(self._add_pass1)
        r1.addWidget(QLabel("确认")); r1.addWidget(self._add_pass2)
        r1.addWidget(self._add_sudo)
        r1.addStretch()
        lay.addLayout(r1)

        self._del_check = QCheckBox("删除用户")
        self._del_user = QLineEdit(); self._del_user.setPlaceholderText("用户名")
        self._del_sudo = QCheckBox("sudo")

        r2 = QHBoxLayout()
        r2.addWidget(self._del_check)
        r2.addWidget(QLabel("用户名")); r2.addWidget(self._del_user)
        r2.addWidget(self._del_sudo)
        r2.addStretch()
        lay.addLayout(r2)

        bar, excel_label = self._create_import_bar()
        lay.addLayout(bar)
        table = self._create_server_table()
        lay.addWidget(table, 1)
        self._tab_pages.append({'table': table, 'excel_label': excel_label, 'excel_path': ''})
        return w

    def _create_ssh_tab(self):
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(4)
        lay.setContentsMargins(20, 10, 20, 8)

        r1 = QHBoxLayout()
        self._port_check = QCheckBox("修改 SSH 端口")
        self._port_input = QLineEdit("22")
        self._port_input.setFixedWidth(70)
        r1.addWidget(self._port_check)
        r1.addWidget(self._port_input)
        r1.addStretch()
        lay.addLayout(r1)

        r2 = QHBoxLayout()
        self._root_check = QCheckBox("禁止 root 远程登录")
        self._root_combo = QComboBox()
        self._root_combo.addItems(["是（禁止）", "否（允许）"])
        r2.addWidget(self._root_check)
        r2.addWidget(QLabel("设置"))
        r2.addWidget(self._root_combo)
        r2.addStretch()
        lay.addLayout(r2)

        bar, excel_label = self._create_import_bar()
        lay.addLayout(bar)
        table = self._create_server_table()
        lay.addWidget(table, 1)
        self._tab_pages.append({'table': table, 'excel_label': excel_label, 'excel_path': ''})
        return w

    def _create_userlist_tab(self):
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(4)
        lay.setContentsMargins(20, 10, 20, 8)

        bar, excel_label = self._create_import_bar()
        lay.addLayout(bar)
        table = self._create_server_table()
        lay.addWidget(table, 1)
        self._tab_pages.append({'table': table, 'excel_label': excel_label, 'excel_path': ''})
        return w

    def _create_security_tab(self):
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(4)
        lay.setContentsMargins(20, 10, 20, 8)

        r1 = QHBoxLayout()
        r1.addWidget(QLabel("并发数（1-10）"))
        self._concurrency_input = QLineEdit("1")
        self._concurrency_input.setFixedWidth(60)
        r1.addWidget(self._concurrency_input)
        r1.addStretch()
        lay.addLayout(r1)

        bar, excel_label = self._create_import_bar()
        lay.addLayout(bar)
        table = self._create_server_table()
        lay.addWidget(table, 1)
        self._tab_pages.append({'table': table, 'excel_label': excel_label, 'excel_path': ''})
        return w

    # ===================== Tab 切换 =====================

    def _on_tab_changed(self, idx):
        if idx < len(self._tab_pages):
            d = self._tab_pages[idx]
            self._table = d['table']
            self._excel_label = d['excel_label']
            self._excel_path = d['excel_path']

    # ===================== 日志操作 =====================

    def _copy_log(self):
        cb = QApplication.clipboard()
        text = self._log_box.toPlainText()
        if text:
            cb.setText(text)
            self._log("日志已复制到剪贴板")

    def _clear_log(self):
        self._log_box.clear()

    # ===================== 表格操作 =====================

    def _on_header_check_clicked(self, idx, table):
        if idx != COL_CHECK or table.rowCount() == 0:
            return
        first = table.item(0, COL_CHECK)
        if not first:
            return
        all_checked = all(
            table.item(r, COL_CHECK) and table.item(r, COL_CHECK).checkState() == Qt.Checked
            for r in range(table.rowCount()))
        new_state = Qt.Unchecked if all_checked else Qt.Checked
        for row in range(table.rowCount()):
            item = table.item(row, COL_CHECK)
            if item:
                item.setCheckState(new_state)
        header = table.horizontalHeader()
        if isinstance(header, _SelectAllHeader):
            header.set_checked(new_state == Qt.Checked)

    def _import_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择 Excel 文件", "", "Excel (*.xlsx)")
        if not path:
            return
        idx = self._tabs.currentIndex()
        d = self._tab_pages[idx]
        d['excel_path'] = path
        d['excel_label'].setText(os.path.basename(path))
        self._excel_path = path
        self._excel_label = d['excel_label']
        try:
            tasks = read_excel_tasks(path)
        except Exception as e:
            QMessageBox.critical(self, "错误", f"读取 Excel 失败: {e}")
            return
        self._table = d['table']
        self._populate_table(tasks)
        self._log(f"已导入 {len(tasks)} 条记录")

    def _add_manual_row(self):
        dialog = _ManualAddDialog(self)
        if dialog.exec() != QDialog.Accepted:
            return
        host, port, user, pwd = dialog.get_result()
        row = self._table.rowCount()
        self._table.insertRow(row)
        self._table.setRowHeight(row, 28)
        chk = QTableWidgetItem()
        chk.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
        chk.setCheckState(Qt.Checked)
        self._table.setItem(row, COL_CHECK, chk)
        self._table.setItem(row, COL_HOST, QTableWidgetItem(host))
        p_item = QTableWidgetItem(str(port)); p_item.setTextAlignment(Qt.AlignCenter)
        self._table.setItem(row, COL_PORT, p_item)
        u_item = QTableWidgetItem(user); u_item.setTextAlignment(Qt.AlignCenter)
        self._table.setItem(row, COL_USER, u_item)
        self._table.setItem(row, COL_PWD, QTableWidgetItem(pwd))
        s_item = QTableWidgetItem("未检测"); s_item.setTextAlignment(Qt.AlignCenter)
        self._table.setItem(row, COL_STATUS, s_item)
        self._table.scrollToBottom()

    def _populate_table(self, tasks):
        self._table.setRowCount(0)
        for host, port, user, pwd in tasks:
            row = self._table.rowCount()
            self._table.insertRow(row)
            self._table.setRowHeight(row, 28)
            chk = QTableWidgetItem()
            chk.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            chk.setCheckState(Qt.Checked)
            self._table.setItem(row, COL_CHECK, chk)
            self._table.setItem(row, COL_HOST, QTableWidgetItem(host))
            p_item = QTableWidgetItem(str(port)); p_item.setTextAlignment(Qt.AlignCenter)
            self._table.setItem(row, COL_PORT, p_item)
            u_item = QTableWidgetItem(user); u_item.setTextAlignment(Qt.AlignCenter)
            self._table.setItem(row, COL_USER, u_item)
            self._table.setItem(row, COL_PWD, QTableWidgetItem(pwd))
            s_item = QTableWidgetItem("未检测"); s_item.setTextAlignment(Qt.AlignCenter)
            self._table.setItem(row, COL_STATUS, s_item)

    def _show_table_context_menu(self, pos, table):
        item = table.itemAt(pos)
        if not item:
            return
        menu = QMenu(self)
        act_del = menu.addAction("删除行")
        act = menu.exec(table.viewport().mapToGlobal(pos))
        if act == act_del:
            row = item.row()
            table.removeRow(row)
            if table.rowCount() == 0:
                header = table.horizontalHeader()
                if isinstance(header, _SelectAllHeader):
                    header.set_checked(False)

    def _delete_checked(self):
        table = self._table
        if table.rowCount() == 0:
            return
        removed = 0
        for row in range(table.rowCount() - 1, -1, -1):
            item = table.item(row, COL_CHECK)
            if item and item.checkState() == Qt.Checked:
                table.removeRow(row)
                removed += 1
        if removed:
            header = table.horizontalHeader()
            if isinstance(header, _SelectAllHeader):
                header.set_checked(False)
            self._log(f"已删除 {removed} 条记录")
        else:
            self._log("没有勾选的记录", "WARNING")

    # 跨线程更新状态信号的连接
    _status_signal = Signal(int, str)

    def _check_ssh_status(self):
        """后台并发检测表格所有行的 SSH 连通性"""
        table = self._table
        if table.rowCount() == 0:
            return

        tasks = []
        for row in range(table.rowCount()):
            host = table.item(row, COL_HOST).text().strip() if table.item(row, COL_HOST) else ""
            if not host:
                continue
            port_str = table.item(row, COL_PORT).text().strip() if table.item(row, COL_PORT) else "22"
            user = table.item(row, COL_USER).text().strip() if table.item(row, COL_USER) else ""
            pwd = table.item(row, COL_PWD).text().strip() if table.item(row, COL_PWD) else ""
            try:
                port = int(port_str)
            except ValueError:
                port = 22
            s_item = QTableWidgetItem("⏳ 检测中"); s_item.setTextAlignment(Qt.AlignCenter)
            table.setItem(row, COL_STATUS, s_item)
            tasks.append((row, host, port, user, pwd))

        if not tasks:
            return

        self._log(f"开始检测 {len(tasks)} 台服务器 SSH 状态...")

        def _run():
            try:
                with ThreadPoolExecutor(max_workers=10) as ex:
                    futs = {ex.submit(ssh_test_login, host, user, pwd, port): row
                            for row, host, port, user, pwd in tasks}
                    for f in as_completed(futs):
                        row = futs[f]
                        try:
                            f.result()
                            self._status_signal.emit(row, "✅ 正常")
                        except Exception:
                            self._status_signal.emit(row, "❌ 不通")
                self._status_signal.emit(-1, "__done__")
            except Exception:
                tb = traceback.format_exc()
                self._status_signal.emit(-1, f"__error__\n{tb}")

        threading.Thread(target=_run, daemon=True).start()

    def _on_status_update(self, row, text):
        if row == -1 and text == "__done__":
            self._log("✅ SSH 状态检测完成", "OK")
            return
        if row == -1 and text.startswith("__error__"):
            self._log("❌ SSH 状态检测异常", "ERROR")
            self._log(text.replace("__error__\n", ""), "ERROR")
            return
        s_item = QTableWidgetItem(text); s_item.setTextAlignment(Qt.AlignCenter)
        self._table.setItem(row, COL_STATUS, s_item)

    def _toggle_select_all(self):
        table = self._table
        if table.rowCount() == 0:
            return
        first_state = table.item(0, COL_CHECK).checkState()
        new_state = Qt.Unchecked if first_state == Qt.Checked else Qt.Checked
        for row in range(table.rowCount()):
            item = table.item(row, COL_CHECK)
            if item:
                item.setCheckState(new_state)
        header = table.horizontalHeader()
        if isinstance(header, _SelectAllHeader):
            header.set_checked(new_state == Qt.Checked)

    def _get_checked_tasks(self):
        tasks = []
        for row in range(self._table.rowCount()):
            chk = self._table.item(row, COL_CHECK)
            if chk and chk.checkState() == Qt.Checked:
                host = self._table.item(row, COL_HOST).text().strip() if self._table.item(row, COL_HOST) else ""
                port_str = self._table.item(row, COL_PORT).text().strip() if self._table.item(row, COL_PORT) else "22"
                user = self._table.item(row, COL_USER).text().strip() if self._table.item(row, COL_USER) else ""
                pwd = self._table.item(row, COL_PWD).text().strip() if self._table.item(row, COL_PWD) else ""
                if not host:
                    continue
                try:
                    port = int(port_str)
                except ValueError:
                    port = 22
                tasks.append((host, port, user, pwd))
        return tasks

    # ===================== 执行逻辑 =====================

    def _log(self, msg, level="INFO"):
        """增强日志：带时间戳、颜色分级、支持多行（如 traceback）"""
        ts = datetime.now().strftime("%H:%M:%S")
        colors = {
            "DEBUG": "#636e72",
            "INFO": "#2d3436",
            "OK": "#27ae60",
            "WARNING": "#e17055",
            "ERROR": "#d63031",
            "CRITICAL": "#d63031",
        }
        color = colors.get(level, "#2d3436")
        prefix_map = {"OK": "", "WARNING": "⚠️ ", "ERROR": "", "CRITICAL": "💥"}
        prefix = prefix_map.get(level, "")

        for line in msg.split("\n"):
            formatted = (
                f'<span style="color:{color};white-space:pre-wrap;">'
                f"[{ts}] {prefix} {line}</span>"
            )
            self._log_box.append(formatted)
        self._log_box.ensureCursorVisible()
        getattr(logger, level.lower(), logger.info)(msg)

    def _auto_log(self, msg):
        """根据消息内容自动判断级别后输出到日志"""
        if "❌" in msg or "失败" in msg or "错误" in msg or "Error" in msg or "exception" in msg.lower():
            self._log(msg, "ERROR")
        elif "⚠️" in msg or "跳过" in msg or "取消" in msg:
            self._log(msg, "WARNING")
        elif "✅" in msg or "完成" in msg or "成功" in msg or "正常" in msg:
            self._log(msg, "OK")
        else:
            self._log(msg, "INFO")

    def _start(self):
        tasks = self._get_checked_tasks()
        if not tasks:
            QMessageBox.critical(self, "错误", "没有勾选的服务器，请先导入 Excel 或手动添加")
            return

        tab_idx = self._tabs.currentIndex()
        self._cancel_flag = _CancelFlag()

        if tab_idx == 0:  # SSH配置修改
            sshd_options = {}
            if self._port_check.isChecked():
                try:
                    port = int(self._port_input.text())
                    if not (1 <= port <= 65535): raise ValueError
                    sshd_options["Port"] = port
                except ValueError:
                    QMessageBox.critical(self, "错误", "SSH 端口不合法"); return
            if self._root_check.isChecked():
                sshd_options["PermitRootLogin"] = (
                    "no" if self._root_combo.currentText().startswith("是") else "yes")
            if not sshd_options:
                QMessageBox.critical(self, "错误", "请至少选择一个配置项"); return
            self._worker = _SSHConfigWorker(tasks, sshd_options, self._cancel_flag)

        elif tab_idx == 1:  # 用户新增删除
            if not self._add_check.isChecked() and not self._del_check.isChecked():
                QMessageBox.critical(self, "错误", "至少选择一个操作"); return
            if self._add_check.isChecked():
                if not self._add_user.text():
                    QMessageBox.critical(self, "错误", "请输入新增用户名"); return
                if self._add_pass1.text() != self._add_pass2.text():
                    QMessageBox.critical(self, "错误", "密码不一致"); return
            if self._del_check.isChecked():
                if not self._del_user.text():
                    QMessageBox.critical(self, "错误", "请输入删除用户名"); return
            add_cfg = {"enable": False}; del_cfg = {"enable": False}
            if self._add_check.isChecked():
                add_cfg = {"enable": True, "user": self._add_user.text(),
                           "pass": self._add_pass1.text(), "sudo": self._add_sudo.isChecked()}
            if self._del_check.isChecked():
                del_cfg = {"enable": True, "user": self._del_user.text(),
                           "sudo": self._del_sudo.isChecked()}
            self._worker = _UserWorker(tasks, add_cfg, del_cfg, self._cancel_flag)

        elif tab_idx == 2:  # 获取用户列表
            out_dir = os.path.join(THIS_DIR, "output")
            os.makedirs(out_dir, exist_ok=True)
            out_path = os.path.join(out_dir,
                f"获取用户列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            self._worker = _UserListWorker(tasks, out_path, self._cancel_flag)

        elif tab_idx == 3:  # 安全基线扫描
            try:
                n = int(self._concurrency_input.text())
                if n < 1 or n > 10: raise ValueError
            except ValueError:
                QMessageBox.critical(self, "错误", "并发数应为 1-10"); return
            out_dir = os.path.join(THIS_DIR, "output")
            os.makedirs(out_dir, exist_ok=True)
            xl_out = os.path.join(out_dir,
                f"安全基线检查_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            self._worker = _SecurityWorker(tasks, xl_out, n, self._cancel_flag)

        self._run_btn.setEnabled(False)
        self._cancel_btn.setEnabled(True)

        self._thread = QThread()
        self._worker.moveToThread(self._thread)
        self._thread.started.connect(self._worker.run)
        self._worker.log_signal.connect(self._auto_log)
        self._worker.finished_signal.connect(self._on_finished)
        self._worker.finished_signal.connect(self._thread.quit)
        self._worker.finished_signal.connect(self._worker.deleteLater)
        self._thread.finished.connect(self._thread.deleteLater)
        self._thread.start()
        self._log(f"=== 开始 [{self._tabs.tabText(tab_idx)}]，共 {len(tasks)} 台 ===", "INFO")

    def _cancel(self):
        if self._cancel_flag:
            self._cancel_flag.cancelled = True
        self._log("=== 用户取消，正在停止... ===", "WARNING")

    def _on_finished(self):
        self._log("✅ === 任务完成 ===", "OK")
        self._run_btn.setEnabled(True)
        self._cancel_btn.setEnabled(False)
        self._worker = None
        QMessageBox.information(self, "完成", "所有服务器已处理完毕")


# ================= 主入口 =================
if __name__ == "__main__":
    # 全局异常钩子：捕获未处理的异常并写日志
    def _global_excepthook(exc_type, exc_value, exc_tb):
        if issubclass(exc_type, KeyboardInterrupt):
            sys.__excepthook__(exc_type, exc_value, exc_tb)
            return
        msg = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
        logging.getLogger(__name__).critical(f"未捕获的全局异常:\n{msg}")
        print(msg, file=sys.stderr)

    sys.excepthook = _global_excepthook
    threading.excepthook = lambda args: _global_excepthook(
        args.exc_type, args.exc_value, args.exc_traceback)

    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s [%(levelname)s] %(message)s",
                        handlers=[logging.StreamHandler()])
    app = QApplication(sys.argv)
    win = LinuxToolsGUI()
    win.show()
    sys.exit(app.exec())
